from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.dashboard_runtime import resolve_snapshot_workbook_path
from app.field_roles import unique_headers
from app.ingest_excel import parse_contracted_vs_actual_xlsx
from app.settings import settings


@dataclass(frozen=True)
class RawTable:
    name: str
    headers: list[str]
    rows: list[dict[str, Any]]


def _coerce_scalar(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if value in (None, ""):
        return None
    if isinstance(value, (int, float, bool)):
        return value
    text = str(value).strip()
    return text or None


def _storage_dir() -> Path:
    root = Path(settings.storage_dir).resolve() / "raw_snapshots"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _snapshot_dir(snapshot_id: int) -> Path:
    target = _storage_dir() / str(snapshot_id)
    target.mkdir(parents=True, exist_ok=True)
    return target


def _generic_sheet_tables(path: Path) -> list[RawTable]:
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    tables: list[RawTable] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row = None
        header_values: list[str] = []
        for rr in range(1, min(ws.max_row, 12) + 1):
            values = [str(value or "").strip() for value in next(ws.iter_rows(min_row=rr, max_row=rr, min_col=1, max_col=ws.max_column, values_only=True))]
            non_empty = [value for value in values if value]
            if len(non_empty) >= 2:
                header_row = rr
                header_values = values
                break
        if header_row is None:
            continue
        included_columns, headers = unique_headers(header_values)
        rows: list[dict[str, Any]] = []
        for values in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, min_col=1, max_col=ws.max_column, values_only=True):
            row = {
                headers[pos]: _coerce_scalar(values[col_idx]) if col_idx < len(values) else None
                for pos, col_idx in enumerate(included_columns)
            }
            if any(value not in (None, "") for value in row.values()):
                rows.append(row)
        tables.append(RawTable(name=sheet_name, headers=headers, rows=rows))
    return tables


def extract_raw_tables_from_workbook(path: Path, workbook_type: str) -> list[RawTable]:
    if workbook_type == "contracted_actual_v1":
        parsed = parse_contracted_vs_actual_xlsx(path)
        client_rows = [
            {
                "client_name": row.name,
                "client_id": row.external_id,
                "client_type": row.client_type,
                "contracted_ind": row.contracted_ind,
                "contracted_bus": row.contracted_bus,
                "contracted_total": row.contracted_total,
                "received_ind": row.received_ind,
                "received_bus": row.received_bus,
                "received_total": row.received_total,
                "pending_ind": row.pending_ind,
                "pending_bus": row.pending_bus,
                "pending_total": row.pending_total,
                "receipt_rate": (row.received_total / row.contracted_total) if row.contracted_total else None,
            }
            for row in parsed.clients
        ]
        staff_rows = [
            {
                "staff_name": row.name,
                "staff_id": row.external_id,
                "staff_type": row.staff_type,
                "received_ind": row.received_ind,
                "received_bus": row.received_bus,
                "received_total": row.received_total,
            }
            for row in parsed.staff
        ]
        tables = [
            RawTable(name="clients", headers=list(client_rows[0].keys()) if client_rows else [], rows=client_rows),
        ]
        if staff_rows:
            tables.append(RawTable(name="staff", headers=list(staff_rows[0].keys()), rows=staff_rows))
        return tables
    return _generic_sheet_tables(path)


def persist_snapshot_raw_tables(snapshot_id: int, workbook_type: str, source_filename: str, tables: list[RawTable]) -> Path:
    target = _snapshot_dir(snapshot_id)
    manifest = {
        "snapshot_id": snapshot_id,
        "workbook_type": workbook_type,
        "source_filename": source_filename,
        "tables": [],
    }
    for table in tables:
        filename = f"{table.name}.json"
        table_path = target / filename
        payload = {"name": table.name, "headers": table.headers, "rows": table.rows}
        table_path.write_text(json.dumps(payload, ensure_ascii=True), encoding="utf-8")
        manifest["tables"].append({"name": table.name, "filename": filename, "row_count": len(table.rows), "column_count": len(table.headers)})
    manifest_path = target / "manifest.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=True, indent=2), encoding="utf-8")
    return manifest_path


def load_snapshot_raw_tables(snapshot_id: int) -> list[RawTable]:
    target = _snapshot_dir(snapshot_id)
    manifest_path = target / "manifest.json"
    if not manifest_path.exists():
        return []
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    tables: list[RawTable] = []
    for entry in manifest.get("tables", []):
        payload = json.loads((target / entry["filename"]).read_text(encoding="utf-8"))
        tables.append(RawTable(name=payload["name"], headers=list(payload.get("headers") or []), rows=list(payload.get("rows") or [])))
    return tables


def load_or_extract_snapshot_raw_tables(snapshot, workbook_type: str) -> list[RawTable]:
    tables = load_snapshot_raw_tables(snapshot.id)
    if tables:
        return tables
    path = resolve_snapshot_workbook_path(snapshot)
    if path is None:
        return []
    tables = extract_raw_tables_from_workbook(path, workbook_type)
    if tables:
        persist_snapshot_raw_tables(snapshot.id, workbook_type, snapshot.source_filename, tables)
    return tables


def load_tables_from_path(path: Path, workbook_type: str) -> list[RawTable]:
    return extract_raw_tables_from_workbook(path, workbook_type)


def raw_table_manifest(snapshot_id: int) -> dict[str, Any] | None:
    manifest_path = _snapshot_dir(snapshot_id) / "manifest.json"
    if not manifest_path.exists():
        return None
    return json.loads(manifest_path.read_text(encoding="utf-8"))
