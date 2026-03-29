from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


DATE_RE = re.compile(r"received\s+as\s+of\s+(\d{1,2})/(\d{1,2})", re.IGNORECASE)


def _safe_int(v) -> int:
    if v is None or v == "":
        return 0
    try:
        return int(v)
    except Exception:
        try:
            return int(float(v))
        except Exception:
            return 0


def _normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def _suggest_pii_type(header_label: str) -> str | None:
    normalized = _normalize_header(header_label)
    if "email" in normalized:
        return "email"
    if "phone" in normalized or "mobile" in normalized or "fax" in normalized:
        return "phone"
    if "address" in normalized or "street" in normalized or "city" in normalized or "zip" in normalized:
        return "address"
    if normalized in {"client id", "staff id", "employee id", "tax id", "ssn", "ein", "id"} or normalized.endswith(" id"):
        return "identifier"
    if "name" in normalized:
        return "name"
    return None


def _row_values(ws, row_idx: int, max_cols: int) -> list[str]:
    return [str(cell or "").strip() for cell in next(ws.iter_rows(min_row=row_idx, max_row=row_idx, min_col=1, max_col=max_cols, values_only=True))]


def _single_value(ws, row_idx: int, col_idx: int):
    return next(ws.iter_rows(min_row=row_idx, max_row=row_idx, min_col=col_idx, max_col=col_idx, values_only=True))[0]


def _looks_like_contracted_layout(ws) -> bool:
    if ws.max_row < 3 or ws.max_column < 6:
        return False
    row1 = _row_values(ws, 1, min(ws.max_column, 12))
    row2 = _row_values(ws, 2, min(ws.max_column, 12))
    return (
        len(row1) >= 3
        and row1[0].lower() == "client name"
        and row1[1].lower() == "client id"
        and row1[2].lower() == "client type"
        and any(value.lower() == "ind" for value in row2)
        and any(value.lower() == "bus" for value in row2)
        and any(value.lower() == "tot" for value in row2)
    )


@dataclass(frozen=True)
class ParsedClientRow:
    name: str
    external_id: str
    client_type: str
    contracted_ind: int
    contracted_bus: int
    contracted_total: int
    received_ind: int
    received_bus: int
    received_total: int
    pending_ind: int
    pending_bus: int
    pending_total: int


@dataclass(frozen=True)
class ParsedStaffRow:
    name: str
    external_id: str
    staff_type: str
    received_ind: int
    received_bus: int
    received_total: int


@dataclass(frozen=True)
class ParsedWorkbook:
    as_of_date: date
    clients: list[ParsedClientRow]
    staff: list[ParsedStaffRow]


@dataclass(frozen=True)
class PreviewHeaderField:
    column: str
    header_label: str
    sample_value: str | None
    suggested_pii_type: str | None


@dataclass(frozen=True)
class PreviewSection:
    section_key: str
    section_label: str
    header_row: int
    headers: list[PreviewHeaderField]


@dataclass(frozen=True)
class PreviewSheet:
    sheet_name: str
    sections: list[PreviewSection]


def infer_as_of_date(ws) -> date:
    """
    The sample sheet has header: 'Received as of 03/09' (month/day). We interpret it
    as current year by default; if the workbook has a year indicator we can refine later.
    """
    # scan first row for 'Received as of'
    for c in range(1, 20):
        v = ws.cell(1, c).value
        if isinstance(v, str):
            m = DATE_RE.search(v)
            if m:
                month = int(m.group(1))
                day = int(m.group(2))
                year = date.today().year
                return date(year, month, day)
    return date.today()


def preview_contracted_vs_actual_xlsx(path: str | Path) -> list[PreviewSheet]:
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    out: list[PreviewSheet] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sections: list[PreviewSection] = []

        # Contracted-vs-actual workbook: row 1 + row 2 form the client header.
        is_contracted_layout = _looks_like_contracted_layout(ws)
        if is_contracted_layout:
            headers: list[PreviewHeaderField] = []
            current_primary = ""
            for col in range(1, min(ws.max_column, 12) + 1):
                primary = str(_single_value(ws, 1, col) or "").strip()
                if primary:
                    current_primary = primary
                elif col >= 4:
                    primary = current_primary
                secondary = str(_single_value(ws, 2, col) or "").strip() if ws.max_row >= 2 else ""
                if primary and secondary:
                    header_label = f"{primary} - {secondary}"
                else:
                    header_label = primary or secondary
                if not header_label:
                    continue
                sample_value = _single_value(ws, 3, col)
                headers.append(
                    PreviewHeaderField(
                        column=get_column_letter(col),
                        header_label=header_label,
                        sample_value=str(sample_value).strip() if sample_value not in (None, "") else None,
                        suggested_pii_type=_suggest_pii_type(header_label),
                    )
                )
            if headers:
                sections.append(
                    PreviewSection(
                        section_key="clients",
                        section_label="Client Table",
                        header_row=1,
                        headers=headers,
                    )
                )
        else:
            # Generic workbook variant: use the first non-empty row as the header row.
            client_header_row = None
            scan_limit = min(ws.max_row, 10)
            for rr, values in enumerate(
                ws.iter_rows(min_row=1, max_row=scan_limit, min_col=1, max_col=ws.max_column, values_only=True),
                start=1,
            ):
                row_values = [str(value or "").strip() for value in values]
                non_empty = [value for value in row_values if value]
                if len(non_empty) >= 3:
                    client_header_row = rr
                    break
            if client_header_row:
                headers: list[PreviewHeaderField] = []
                sample_row = client_header_row + 1 if client_header_row < ws.max_row else client_header_row
                header_values = next(ws.iter_rows(min_row=client_header_row, max_row=client_header_row, min_col=1, max_col=ws.max_column, values_only=True))
                sample_values = next(ws.iter_rows(min_row=sample_row, max_row=sample_row, min_col=1, max_col=ws.max_column, values_only=True))
                for col, raw_header in enumerate(header_values, start=1):
                    header_label = str(raw_header or "").strip()
                    if not header_label:
                        continue
                    sample_value = sample_values[col - 1] if col - 1 < len(sample_values) else None
                    headers.append(
                        PreviewHeaderField(
                            column=get_column_letter(col),
                            header_label=header_label,
                            sample_value=str(sample_value).strip() if sample_value not in (None, "") else None,
                            suggested_pii_type=_suggest_pii_type(header_label),
                        )
                    )
                if headers:
                    sections.append(
                        PreviewSection(
                            section_key="clients",
                            section_label="Client Table",
                            header_row=client_header_row,
                            headers=headers,
                        )
                    )

        staff_header_row = None
        if is_contracted_layout:
            for rr, values in enumerate(
                ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=min(ws.max_column, 4), values_only=True),
                start=1,
            ):
                a = values[0] if len(values) >= 1 else None
                d = values[3] if len(values) >= 4 else None
                if isinstance(a, str) and isinstance(d, str):
                    if a.strip().lower() == "client name" and d.strip().lower().startswith("received"):
                        staff_header_row = rr
                        break

        if staff_header_row:
            headers: list[PreviewHeaderField] = []
            current_primary = ""
            for col in range(1, 7):
                primary = str(_single_value(ws, staff_header_row, col) or "").strip()
                if primary:
                    current_primary = primary
                elif col >= 4:
                    primary = current_primary
                secondary = str(_single_value(ws, staff_header_row + 1, col) or "").strip() if staff_header_row + 1 <= ws.max_row else ""
                header_label = f"{primary} - {secondary}".strip(" -") if secondary else primary
                if not header_label:
                    continue
                sample_value = _single_value(ws, staff_header_row + 2, col) if staff_header_row + 2 <= ws.max_row else None
                headers.append(
                    PreviewHeaderField(
                        column=get_column_letter(col),
                        header_label=header_label,
                        sample_value=str(sample_value).strip() if sample_value not in (None, "") else None,
                        suggested_pii_type=_suggest_pii_type(header_label),
                    )
                )
            if headers:
                sections.append(
                    PreviewSection(
                        section_key="staff",
                        section_label="Staff Table",
                        header_row=staff_header_row,
                        headers=headers,
                    )
                )

        if sections:
            out.append(PreviewSheet(sheet_name=sheet_name, sections=sections))

    return out


def parse_contracted_vs_actual_xlsx(path: str | Path) -> ParsedWorkbook:
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    if "Est vs Actual-2026" in wb.sheetnames:
        ws = wb["Est vs Actual-2026"]
    else:
        ws = wb[wb.sheetnames[0]]

    if not _looks_like_contracted_layout(ws):
        raise ValueError("Workbook does not match the supported Contracted vs Actual layout")

    as_of = infer_as_of_date(ws)

    clients: list[ParsedClientRow] = []
    staff: list[ParsedStaffRow] = []

    # client rows appear from row 3 until the first "Total" row at column A.
    r = 3
    while r <= ws.max_row:
        first = ws.cell(r, 1).value
        if first is None:
            r += 1
            continue
        if isinstance(first, str) and first.strip().lower() == "total":
            break
        # A-L: name, id, type, contracted ind/bus/tot, received ind/bus/tot, pending ind/bus/tot
        name = str(ws.cell(r, 1).value).strip()
        ext_id = str(ws.cell(r, 2).value or "").strip()
        ctype = str(ws.cell(r, 3).value or "CPA").strip()
        con_ind = _safe_int(ws.cell(r, 4).value)
        con_bus = _safe_int(ws.cell(r, 5).value)
        con_tot = _safe_int(ws.cell(r, 6).value) or (con_ind + con_bus)
        rec_ind = _safe_int(ws.cell(r, 7).value)
        rec_bus = _safe_int(ws.cell(r, 8).value)
        rec_tot = _safe_int(ws.cell(r, 9).value) or (rec_ind + rec_bus)
        pend_ind = _safe_int(ws.cell(r, 10).value)
        pend_bus = _safe_int(ws.cell(r, 11).value)
        pend_tot = _safe_int(ws.cell(r, 12).value) or (con_tot - rec_tot)

        clients.append(
            ParsedClientRow(
                name=name,
                external_id=ext_id,
                client_type=ctype,
                contracted_ind=con_ind,
                contracted_bus=con_bus,
                contracted_total=con_tot,
                received_ind=rec_ind,
                received_bus=rec_bus,
                received_total=rec_tot,
                pending_ind=pend_ind,
                pending_bus=pend_bus,
                pending_total=pend_tot,
            )
        )
        r += 1

    # find staff section start (row where A == 'Client Name' and D == 'Received Till date')
    staff_start = None
    for rr in range(1, ws.max_row + 1):
        a = ws.cell(rr, 1).value
        d = ws.cell(rr, 4).value
        if isinstance(a, str) and isinstance(d, str):
            if a.strip().lower() == "client name" and d.strip().lower().startswith("received"):
                staff_start = rr + 2  # header + subheader
                break

    if staff_start:
        rr = staff_start
        while rr <= ws.max_row:
            a = ws.cell(rr, 1).value
            if a is None:
                rr += 1
                continue
            if isinstance(a, str) and a.strip().lower() == "total":
                break
            name = str(ws.cell(rr, 1).value).strip()
            ext_id = str(ws.cell(rr, 2).value or "").strip()
            stype = str(ws.cell(rr, 3).value or "").strip()
            rec_ind = _safe_int(ws.cell(rr, 4).value)
            rec_bus = _safe_int(ws.cell(rr, 5).value)
            rec_tot = _safe_int(ws.cell(rr, 6).value) or (rec_ind + rec_bus)
            staff.append(
                ParsedStaffRow(
                    name=name,
                    external_id=ext_id,
                    staff_type=stype,
                    received_ind=rec_ind,
                    received_bus=rec_bus,
                    received_total=rec_tot,
                )
            )
            rr += 1

    return ParsedWorkbook(as_of_date=as_of, clients=clients, staff=staff)
