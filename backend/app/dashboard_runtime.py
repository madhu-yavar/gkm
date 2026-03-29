from __future__ import annotations

from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any
import re

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app import models
from app.field_roles import FieldRoleProfile, classify_table_fields, coerce_measure_value, dimension_headers, measure_headers, unique_headers
from app.settings import settings


def _storage_candidates(snapshot: models.Snapshot) -> list[Path]:
    storage = Path(settings.storage_dir).resolve()
    candidates = [
        storage / f"upload_{snapshot.source_filename}",
        storage / f"preview_{snapshot.source_filename}",
    ]
    candidates.extend(sorted(storage.glob(f"upload_*_{snapshot.source_filename}"), reverse=True))
    candidates.extend(sorted(storage.glob(f"preview_*_{snapshot.source_filename}"), reverse=True))
    seen: set[Path] = set()
    unique: list[Path] = []
    for path in candidates:
        if path in seen:
            continue
        seen.add(path)
        unique.append(path)
    return unique


def resolve_snapshot_workbook_path(snapshot: models.Snapshot) -> Path | None:
    for path in _storage_candidates(snapshot):
        if path.exists():
            return path
    return None


def _parse_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if value in (None, ""):
        return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _normalize_status(value: str) -> str:
    return " ".join(str(value or "").strip().split())


def _coerce_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    if text.endswith("%"):
        text = text[:-1]
    try:
        return float(text)
    except ValueError:
        return None


def _detect_header_row(ws) -> tuple[int | None, list[str]]:
    for rr in range(1, min(ws.max_row, 12) + 1):
        values = [
            str(value or "").strip()
            for value in next(
                ws.iter_rows(
                    min_row=rr,
                    max_row=rr,
                    min_col=1,
                    max_col=ws.max_column,
                    values_only=True,
                )
            )
        ]
        non_empty = [value for value in values if value]
        if len(non_empty) >= 2:
            return rr, values
    return None, []


def _sheet_group(name: str) -> tuple[str | None, str | None]:
    match = re.match(r"^([A-Za-z]+)[ _-]+(.+)$", str(name or "").strip())
    if not match:
        return None, None
    return match.group(1), match.group(2)


def _normalize_header_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower()).strip()


def _find_header(headers: list[str], aliases: list[str]) -> str | None:
    normalized_aliases = [_normalize_header_name(alias) for alias in aliases]
    normalized_headers = {header: _normalize_header_name(header) for header in headers}
    for alias in normalized_aliases:
        for header, normalized in normalized_headers.items():
            if normalized == alias:
                return header
    for alias in normalized_aliases:
        for header, normalized in normalized_headers.items():
            if alias in normalized:
                return header
    return None


def _extract_period_group(name: str) -> tuple[int, int] | None:
    text = str(name or "").strip().lower()
    match = re.search(r"\bq([1-4])\s*[- ]?\s*(\d{2})[-/ ]?(\d{2})\b", text)
    if match:
        quarter = int(match.group(1))
        start_year = 2000 + int(match.group(2))
        return start_year, quarter
    return None


def _period_label(name: str) -> str:
    return str(name or "").strip()


def _extract_tower_code(value: Any) -> str | None:
    text = str(value or "").strip().upper()
    if not text or text in {"TOTAL", "SUMMARY"}:
        return None
    match = re.match(r"^(\d{1,2}[A-Z]?)", text)
    if match:
        return match.group(1)
    match = re.match(r"^([A-Z]{1,3}\d{1,2})", text)
    return match.group(1) if match else None


def _preferred_chart_type(preferences: list[str], allowed: list[str], default: str) -> str:
    for chart in preferences:
        if chart in allowed:
            return chart
    return default


def _format_value(value: float) -> str:
    return f"{value:,.0f}" if abs(value) >= 100 else f"{value:,.2f}"


def _build_generic_table_like(headers: list[str], rows: list[dict[str, Any]]):
    return type("TableLike", (), {"headers": headers, "rows": rows})()  # type: ignore[misc]


def _collect_collections_records(raw_tables) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    records: list[dict[str, Any]] = []
    period_totals: list[dict[str, Any]] = []
    for table in raw_tables:
        if str(table.name or "").strip().lower() in {"definitions", "questions"}:
            continue
        if not table.headers or not table.rows:
            continue
        headers = table.headers
        unit_header = _find_header(headers, ["unit", "house", "number", "flat", "tower"])
        owner_header = _find_header(headers, ["owner name", "owner", "customer", "member"])
        dues_header = _find_header(headers, ["total dues", "dues", "balance", "outstanding", "amount due"])
        penalty_header = _find_header(headers, ["accumulated penalty", "penalty", "late fee"])
        count_header = _find_header(headers, ["total count", "count"])
        analysis_scope = (
            "pending_dues"
            if (dues_header and "dues" in _normalize_header_name(dues_header)) or penalty_header
            else "period_ledger"
        )
        if not unit_header:
            profiles = classify_table_fields(_build_generic_table_like(headers, table.rows))
            dimension_candidates = dimension_headers(table, profiles)
            unit_header = _find_header(dimension_candidates, ["unit", "house", "tower"]) or (dimension_candidates[0] if dimension_candidates else None)
        if not dues_header:
            profiles = classify_table_fields(_build_generic_table_like(headers, table.rows))
            measure_candidates = measure_headers(table, profiles)
            dues_header = _find_header(measure_candidates, ["dues", "balance", "amount", "total"]) or (measure_candidates[0] if measure_candidates else None)
        if not unit_header or (not dues_header and not penalty_header):
            continue
        period_key = _extract_period_group(table.name)
        period_label = _period_label(table.name)
        due_total = 0.0
        penalty_total = 0.0
        row_count = 0
        for row in table.rows:
            unit_value = row.get(unit_header)
            unit = str(unit_value or "").strip()
            if not unit or unit.lower() in {"total", "summary"}:
                continue
            tower = _extract_tower_code(unit)
            if not tower:
                continue
            dues = _coerce_number(row.get(dues_header)) if dues_header else None
            penalty = _coerce_number(row.get(penalty_header)) if penalty_header else None
            count_value = _coerce_number(row.get(count_header)) if count_header else None
            if dues is None and penalty is None and count_value is None:
                continue
            owner = str(row.get(owner_header) or "").strip() if owner_header else ""
            record = {
                "sheet_name": table.name,
                "period_label": period_label,
                "period_key": period_key,
                "unit": unit,
                "tower": tower,
                "owner": owner or "Unknown",
                "dues": float(dues or 0.0),
                "penalty": float(penalty or 0.0),
                "count": float(count_value or 0.0),
                "analysis_scope": analysis_scope,
            }
            records.append(record)
            due_total += record["dues"]
            penalty_total += record["penalty"]
            row_count += 1
        if row_count:
            period_totals.append(
                {
                    "sheet_name": table.name,
                    "period_label": period_label,
                    "period_key": period_key,
                    "dues_total": due_total,
                    "penalty_total": penalty_total,
                    "row_count": row_count,
                }
            )
    return records, period_totals


def _ranked_items(records: list[dict[str, Any]], key: str, value_key: str, top_n: int = 5) -> list[dict[str, Any]]:
    totals: dict[str, float] = {}
    for record in records:
        label = str(record.get(key) or "").strip()
        if not label:
            continue
        totals[label] = totals.get(label, 0.0) + float(record.get(value_key) or 0.0)
    ranked = sorted(totals.items(), key=lambda item: item[1], reverse=True)
    total_value = sum(value for _, value in ranked) or 1.0
    return [
        {
            "label": label,
            "value": value,
            "share": value / total_value,
        }
        for label, value in ranked[:top_n]
        if value > 0
    ]


def _ranked_scope_items(records: list[dict[str, Any]], scope: str, key: str, value_key: str, top_n: int = 5) -> list[dict[str, Any]]:
    scoped = [record for record in records if str(record.get("analysis_scope") or "") == scope]
    return _ranked_items(scoped or records, key, value_key, top_n=top_n)


def _period_series(period_totals: list[dict[str, Any]], value_key: str) -> list[dict[str, Any]]:
    ordered = sorted(
        [item for item in period_totals if item.get("period_key") is not None and (item.get(value_key) or 0.0) > 0],
        key=lambda item: item["period_key"],
    )
    return [
        {
            "label": str(item["period_label"]),
            "value": float(item.get(value_key) or 0.0),
            "meta": f"{int(item.get('row_count') or 0)} units",
        }
        for item in ordered
    ]


def _period_gantt(period_totals: list[dict[str, Any]]) -> list[dict[str, Any]]:
    ordered = sorted(
        [item for item in period_totals if item.get("period_key") is not None],
        key=lambda item: item["period_key"],
    )
    return [
        {
            "label": str(item["period_label"]),
            "start": index,
            "end": index + 1,
            "value": float(item.get("dues_total") or 0.0),
        }
        for index, item in enumerate(ordered)
    ]


def _project_next_series_value(series: list[dict[str, Any]]) -> float:
    values = [float(item.get("value") or 0.0) for item in series if float(item.get("value") or 0.0) > 0]
    if not values:
        return 0.0
    if len(values) == 1:
        return values[0]
    slope = values[-1] - values[-2]
    return max(0.0, values[-1] + slope)


def _top_odr_option_items(comparison_groups: list[dict[str, Any]], top_n: int = 5) -> tuple[dict[str, list[dict[str, Any]]], list[str], str | None]:
    options: dict[str, list[dict[str, Any]]] = {}
    ordered_groups = sorted(
        [item for item in comparison_groups if item.get("highest_rate_segments")],
        key=lambda item: str(item.get("group_label") or ""),
    )
    for group in ordered_groups:
        label = str(group.get("group_label") or "").strip()
        if not label:
            continue
        rows = []
        for item in list(group.get("highest_rate_segments") or [])[:top_n]:
            ratio = float(item.get("ratio") or 0.0)
            if ratio <= 0:
                continue
            rows.append(
                {
                    "label": str(item.get("label") or "Unknown"),
                    "value": ratio * 100.0,
                    "meta": f"BC {int(item.get('bad_count') or 0)} / TC {int(item.get('total_count') or 0)}",
                }
            )
        if rows:
            options[label] = rows
    option_labels = list(options.keys())
    selected_option = option_labels[-1] if option_labels else None
    return options, option_labels, selected_option


def _semantic_label(value: Any) -> str:
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, dict):
        for key in ("measure_name", "dimension_name", "entity_name", "name", "column_name", "label", "title", "business_meaning", "description"):
            raw = value.get(key)
            if isinstance(raw, str) and raw.strip():
                return raw.strip()
    return str(value or "").strip()


def _share_items(items: list[dict[str, Any]], label_key: str, value_key: str, top_n: int = 6) -> list[dict[str, Any]]:
    ranked = [
        {
            "label": str(item.get(label_key) or "Unknown"),
            "value": float(item.get(value_key) or 0.0),
        }
        for item in items
        if float(item.get(value_key) or 0.0) > 0
    ]
    ranked = sorted(ranked, key=lambda item: item["value"], reverse=True)[:top_n]
    total = sum(item["value"] for item in ranked) or 1.0
    return [{**item, "share": item["value"] / total} for item in ranked]


def _sheet_grand_total_series(sheet_summaries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "label": str(sheet.get("sheet_name") or "Sheet"),
            "value": float(sheet.get("grand_total") or 0.0),
            "meta": f"{int(sheet.get('row_count') or 0)} rows",
        }
        for sheet in sheet_summaries
        if sheet.get("sheet_kind") == "distribution" and float(sheet.get("grand_total") or 0.0) > 0
    ]


def _configured_tab_keys(dashboard_config: dict[str, Any]) -> set[str]:
    return {
        str(tab.get("key") or "").strip()
        for tab in list(dashboard_config.get("tabs") or [])
        if str(tab.get("key") or "").strip()
    }


def _pick_tab(tab_keys: set[str], *preferred: str) -> str:
    for key in preferred:
        if key in tab_keys:
            return key
    return next(iter(tab_keys), "overview")


def _preferred_trend_chart(preferences: list[str], default: str = "line") -> str:
    for candidate in ("scatter", "gantt", "line"):
        if candidate in preferences:
            return candidate
    return default


def build_adaptive_generic_runtime(snapshot: models.Snapshot, dashboard_config: dict[str, Any]) -> dict[str, Any]:
    from app.raw_data_store import load_or_extract_snapshot_raw_tables

    raw_tables = load_or_extract_snapshot_raw_tables(snapshot, "generic_workbook_v1")
    spec = dict(dashboard_config.get("adaptive_dashboard_spec") or {})
    tab_keys = _configured_tab_keys(dashboard_config) or {"overview", "analysis"}
    chart_preferences = [_semantic_label(item) for item in list(spec.get("chart_preferences") or dashboard_config.get("chart_preferences") or []) if _semantic_label(item)]
    domain = _semantic_label(spec.get("domain"))
    primary_entity = _semantic_label(spec.get("primary_entity") or "entity")
    primary_measure = _semantic_label(spec.get("primary_measure") or "value")
    records, period_totals = _collect_collections_records(raw_tables)
    if domain.lower() in {"", "unknown", "adaptive analytics", "generic analytics"}:
        domain = ""
    if primary_entity.lower() in {"", "entity", "property unit", "property unit id"}:
        primary_entity = ""
    if primary_measure.lower() in {"", "value", "measure"}:
        primary_measure = ""
    if not domain and records:
        domain = "collections analytics"
    if domain == "collections analytics":
        if not primary_entity:
            tower_count = len({record["tower"] for record in records if record.get("tower")})
            primary_entity = "tower" if tower_count else "unit"
        if not primary_measure:
            dues_total = sum(record["dues"] for record in records)
            penalty_total = sum(record["penalty"] for record in records)
            primary_measure = "total dues" if dues_total >= penalty_total else "accumulated penalty"
        top_towers_by_dues = _ranked_scope_items(records, "pending_dues", "tower", "dues", top_n=5)
        top_towers_by_penalty = _ranked_scope_items(records, "pending_dues", "tower", "penalty", top_n=5)
        top_units_by_dues = _ranked_scope_items(records, "pending_dues", "unit", "dues", top_n=8)
        top_owners_by_dues = _ranked_scope_items(records, "pending_dues", "owner", "dues", top_n=8)
        dues_series = _period_series(period_totals, "dues_total")
        penalty_series = _period_series(period_totals, "penalty_total")
        kpis = [
            {"key": "tower_count", "label": "Towers", "value": str(len({record['tower'] for record in records if record.get('tower')})), "meta": "Distinct tower blocks with dues exposure"},
            {"key": "unit_count", "label": "Units", "value": str(len({record['unit'] for record in records if record.get('unit')})), "meta": "Units captured in actionable collections tables"},
            {"key": "total_dues", "label": "Total Dues", "value": _format_value(sum(record["dues"] for record in records)), "meta": "Combined dues across actionable sheets"},
            {"key": "total_penalty", "label": "Total Penalty", "value": _format_value(sum(record["penalty"] for record in records)), "meta": "Accumulated penalty across actionable sheets"},
        ]
        widgets = [
            {
                "key": "top_towers_dues",
                "tab": _pick_tab(tab_keys, "overview"),
                "title": "Top 5 Towers by Total Dues",
                "description": "Highest-exposure towers based on the summed dues captured in the workbook.",
                "chart_type": _preferred_chart_type(chart_preferences, ["pie", "bar", "table"], "bar"),
                "items": top_towers_by_dues,
                "insight": f"The highest-dues tower currently leads the portfolio with {_format_value(top_towers_by_dues[0]['value'])} of dues." if top_towers_by_dues else "No dues-bearing tower records were available for ranking.",
            },
            {
                "key": "top_towers_penalty",
                "tab": _pick_tab(tab_keys, "comparison", "analysis"),
                "title": "Top 5 Towers by Penalty",
                "description": "Penalty-heavy towers that may need targeted collections action.",
                "chart_type": _preferred_chart_type(chart_preferences, ["pie", "bar", "table"], "bar"),
                "items": top_towers_by_penalty,
                "insight": f"Penalty exposure is concentrated in tower {top_towers_by_penalty[0]['label']}." if top_towers_by_penalty else "No penalty-bearing tower records were available for ranking.",
            },
            {
                "key": "owner_exposure",
                "tab": _pick_tab(tab_keys, "comparison", "analysis"),
                "title": "Top Owners by Dues",
                "description": "Owner-wise dues concentration across the actionable sheets.",
                "chart_type": "table",
                "rows": [
                    {"Owner": item["label"], "Dues": _format_value(item["value"]), "Share": f"{item['share'] * 100:.1f}%"}
                    for item in top_owners_by_dues[:5]
                ],
                "insight": "Owner-level concentration can guide targeted follow-up where the same owner appears in multiple high-dues units." if top_owners_by_dues else "No owner-level dues rows were available.",
            },
            {
                "key": "dues_trend",
                "tab": _pick_tab(tab_keys, "trends", "analysis"),
                "title": "Quarter-wise Dues Movement",
                "description": "Total dues or balance exposure across the available quarter sheets.",
                "chart_type": _preferred_trend_chart(chart_preferences, "line"),
                "items": dues_series,
                "insight": "Quarter totals now provide a direct dues trajectory instead of sheet-level review diagnostics." if dues_series else "No quarter-aligned dues totals were available for trend analysis.",
            },
            {
                "key": "trend_snapshot",
                "tab": _pick_tab(tab_keys, "overview"),
                "title": "Dues Trend Snapshot",
                "description": "Quick quarter-wise dues direction surfaced directly on the overview tab.",
                "chart_type": _preferred_trend_chart(chart_preferences, "line"),
                "items": dues_series,
                "insight": "The overview carries the same dues-trend signal used in the detailed trends tab." if dues_series else "No trend series was available for the overview snapshot.",
            },
            {
                "key": "timeline_coverage",
                "tab": _pick_tab(tab_keys, "trends", "comparison", "analysis"),
                "title": "Quarter Coverage Timeline",
                "description": "Workbook periods surfaced for collections analysis.",
                "chart_type": "gantt" if "gantt" in chart_preferences else "table",
                "items": _period_gantt(period_totals),
                "rows": [
                    {"Quarter": item["period_label"], "Dues": _format_value(item["dues_total"]), "Penalty": _format_value(item["penalty_total"]), "Rows": str(item["row_count"])}
                    for item in sorted([entry for entry in period_totals if entry.get("period_key") is not None], key=lambda entry: entry["period_key"])
                ],
                "insight": "The timeline shows which quarters contain usable collections evidence and supports trend comparisons across the period set." if period_totals else "No period-aligned sheets were available for a timeline view.",
            },
            {
                "key": "forecast_outlook",
                "tab": _pick_tab(tab_keys, "trends", "analysis"),
                "title": "Forecast Outlook",
                "description": "Deterministic next-period view based on the aligned dues history.",
                "chart_type": "bar",
                "items": (
                    [
                        {"label": "Current", "value": dues_series[-1]["value"]},
                        {"label": "Projected Next", "value": _project_next_series_value(dues_series)},
                    ]
                    if dues_series else []
                ),
                "insight": "The forecast outlook compares the latest dues level with a deterministic next-period projection derived from recent quarter movement." if dues_series else "Forecast is unavailable because the dues trend series is not populated.",
            },
            {
                "key": "quality_flags",
                "tab": _pick_tab(tab_keys, "quality", "analysis"),
                "title": "Coverage and Modeling Notes",
                "description": "Material coverage constraints for this adaptive dashboard.",
                "chart_type": "table",
                "rows": [
                    {"Check": "Actionable sheets", "Result": str(sum(1 for item in period_totals if item.get("row_count")))},
                    {"Check": "Quarter periods", "Result": str(len([item for item in period_totals if item.get("period_key") is not None]))},
                    {"Check": "Reference sheets excluded", "Result": "Yes"},
                    {"Check": "Requested charts", "Result": ", ".join(chart_preferences) if chart_preferences else "Default adaptive charts"},
                ],
                "insight": "Reference-style invoice sheets were excluded from primary rankings, and the dashboard focuses on actionable collections measures instead.",
            },
        ]
        if "drilldown" in tab_keys:
            widgets.append(
                {
                    "key": "unit_drilldown",
                    "tab": "drilldown",
                    "title": "Top Units by Dues",
                    "description": "Unit-level dues exposure for focused follow-up on specific items.",
                    "chart_type": "table",
                    "rows": [
                        {"Unit": item["label"], "Dues": _format_value(item["value"]), "Share": f"{item['share'] * 100:.1f}%"}
                        for item in top_units_by_dues[:10]
                    ],
                    "insight": "Use this view to inspect the specific units driving tower-level exposure." if top_units_by_dues else "No unit-level dues rows were available for drilldown.",
                }
            )
        return {
            "mode": "adaptive_semantic_runtime",
            "adaptive_dashboard": {
                "domain": domain,
                "primary_entity": primary_entity,
                "primary_measure": primary_measure,
                "chart_preferences": chart_preferences,
                "kpis": kpis,
                "widgets": widgets,
                "supporting_notes": [
                    "Tower codes are derived from unit or house labels such as 18C-18059 and 20A-20052.",
                    "Reference-only sheets are excluded from primary dues and penalty rankings.",
                ],
            },
            "total_sheets": len(raw_tables),
            "tabular_sheet_count": sum(1 for table in raw_tables if table.rows),
            "reference_sheet_count": sum(1 for table in raw_tables if not table.rows),
            "total_rows": sum(len(table.rows) for table in raw_tables),
            "numeric_measure_count": 4,
            "comparison_group_count": len([item for item in period_totals if item.get("period_key") is not None]),
        }

    base = build_generic_workbook_runtime(resolve_snapshot_workbook_path(snapshot)) if resolve_snapshot_workbook_path(snapshot) else {}
    sheet_summaries = [dict(item) for item in list(base.get("sheet_summaries") or []) if item.get("sheet_kind") == "distribution"]
    dominant_sheet = max(
        sheet_summaries,
        key=lambda item: (float(item.get("grand_total") or 0.0), int(item.get("row_count") or 0)),
        default=None,
    )
    if not domain and dominant_sheet:
        headers_lower = {str(header or "").strip().lower() for header in list(dominant_sheet.get("headers") or [])}
        if any(term in headers_lower for term in {"preparer", "reviewer", "client name", "tax payer", "status"}) and (
            any("time" in header or "hour" in header for header in headers_lower)
        ):
            domain = "operations productivity"
    domain = domain or "adaptive analytics"
    comparison_groups = list(base.get("comparison_groups") or [])
    requested_features = [str(item or "").strip() for item in list(spec.get("requested_features") or []) if str(item or "").strip()]
    controls = [str(item or "").strip() for item in list(spec.get("controls") or []) if str(item or "").strip()]
    exposure_items = _share_items(list((dominant_sheet or {}).get("segment_totals") or []), "label", "total", top_n=8)
    measure_mix_items = _share_items(list((dominant_sheet or {}).get("measure_totals") or []), "label", "total", top_n=6)
    duration_mix_items = _share_items(list((dominant_sheet or {}).get("duration_totals") or []), "label", "total", top_n=6)
    sheet_series = _sheet_grand_total_series(sheet_summaries)
    projected_next = _project_next_series_value(sheet_series)
    odr_option_items, odr_options, odr_selected_option = _top_odr_option_items(comparison_groups)
    if not primary_entity and dominant_sheet:
        primary_entity = str(dominant_sheet.get("dimension_header") or "category")
    if not primary_measure and dominant_sheet:
        top_measure = next(iter(list((dominant_sheet or {}).get("measure_totals") or [])), None)
        primary_measure = str((top_measure or {}).get("label") or "total value")
    base["mode"] = "adaptive_semantic_runtime"
    base["adaptive_dashboard"] = {
        "domain": domain,
        "primary_entity": primary_entity,
        "primary_measure": primary_measure,
        "chart_preferences": chart_preferences,
        "kpis": [
            {"key": "sheet_count", "label": "Data Sheets", "value": str(base.get("tabular_sheet_count") or 0), "meta": "Sheets with row-level data"},
            {"key": "row_count", "label": "Rows", "value": _format_value(float(base.get("total_rows") or 0)), "meta": "Visible rows across the workbook"},
            {"key": "measure_count", "label": "Measures", "value": str(base.get("numeric_measure_count") or 0), "meta": "Numeric or duration measures surfaced by field typing"},
            {"key": "dominant_sheet", "label": "Lead Sheet", "value": str((dominant_sheet or {}).get("sheet_name") or "—"), "meta": "Primary source used for adaptive EDA charts"},
        ],
        "widgets": [
            {
                "key": "exposure_ranking",
                "tab": _pick_tab(tab_keys, "overview"),
                "title": f"Top {primary_entity.title() if primary_entity else 'Category'} Exposure",
                "description": "Largest visible segments from the dominant analytical sheet.",
                "chart_type": _preferred_chart_type(chart_preferences, ["pie", "bar", "table"], "bar"),
                "items": exposure_items,
                "insight": f"The dominant sheet '{str((dominant_sheet or {}).get('sheet_name') or 'Unknown')}' is led by {exposure_items[0]['label']} with {_format_value(exposure_items[0]['value'])} of visible exposure." if exposure_items else "No ranked segment exposure was available from the dominant sheet.",
            },
            {
                "key": "overview_trend_snapshot",
                "tab": _pick_tab(tab_keys, "overview"),
                "title": "Volume Snapshot",
                "description": "Quick view of the strongest analytical volume surfaced from the dominant sheet set.",
                "chart_type": _preferred_trend_chart(chart_preferences, "line"),
                "items": sheet_series,
                "insight": "The overview exposes the same sheet-level volume series used for deeper trend analysis." if len(sheet_series) > 1 else "Only one analytical sheet carried enough measure volume for a meaningful trend snapshot.",
            },
            {
                "key": "measure_mix",
                "tab": _pick_tab(tab_keys, "comparison", "analysis"),
                "title": "Measure Composition",
                "description": "How the strongest numeric measures contribute within the dominant sheet.",
                "chart_type": _preferred_chart_type(chart_preferences, ["pie", "bar", "table"], "bar"),
                "items": measure_mix_items or duration_mix_items,
                "insight": f"The dominant sheet is primarily driven by {measure_mix_items[0]['label']}." if measure_mix_items else ("Duration-heavy signals dominate this workbook." if duration_mix_items else "No measure composition was available from the dominant sheet."),
            },
            {
                "key": "sheet_trend",
                "tab": _pick_tab(tab_keys, "trends", "analysis"),
                "title": "Sheet-Level Volume Trend",
                "description": "Grand-total movement across the analytical sheets in workbook order.",
                "chart_type": _preferred_trend_chart(chart_preferences, "line"),
                "items": sheet_series,
                "insight": "The workbook exposes enough sheet-level movement to compare total analytical volume across the visible sheet set." if len(sheet_series) > 1 else "Only one analytical sheet carried enough measure volume for trend comparison.",
            },
            {
                "key": "forecast_outlook",
                "tab": _pick_tab(tab_keys, "trends", "analysis"),
                "title": "Forecast Outlook",
                "description": "Deterministic next-step projection from the sheet-level volume series.",
                "chart_type": "bar",
                "items": (
                    [
                        {"label": "Current", "value": float(sheet_series[-1]["value"])},
                        {"label": "Projected Next", "value": projected_next},
                    ]
                    if sheet_series
                    else []
                ),
                "insight": "The forecast is derived from the recent sheet-level movement and should be treated as directional guidance." if sheet_series else "Forecast is unavailable because the workbook did not expose a usable volume series.",
            },
            {
                "key": "quality_flags",
                "tab": _pick_tab(tab_keys, "quality", "analysis"),
                "title": "Coverage and Modeling Notes",
                "description": "Dominant-sheet context and generic adaptive-modeling checks.",
                "chart_type": "table",
                "rows": [
                    {"Check": "Dominant sheet", "Result": str((dominant_sheet or {}).get("sheet_name") or "—")},
                    {"Check": "Primary dimension", "Result": str((dominant_sheet or {}).get("dimension_header") or "—")},
                    {"Check": "Sheet measures", "Result": ", ".join(str(item.get("label") or "—") for item in list((dominant_sheet or {}).get("measure_totals") or [])[:4]) or "—"},
                    {"Check": "Reference sheets", "Result": str(base.get("reference_sheet_count") or 0)},
                    {"Check": "Requested charts", "Result": ", ".join(chart_preferences) if chart_preferences else "Adaptive defaults"},
                ],
                "insight": "This adaptive fallback is using the strongest typed sheet and its leading measures instead of a domain-specific dashboard family.",
            },
        ]
        + ([
            {
                "key": "odr_top5_month",
                "tab": _pick_tab(tab_keys, "trends"),
                "title": "Top 5 ODR by Month",
                "description": "Select a month and compare the highest ODR pools for that period.",
                "chart_type": "bar",
                "value_format": "percent",
                "options": odr_options,
                "selected_option": odr_selected_option,
                "option_items": odr_option_items,
                "items": odr_option_items.get(odr_selected_option or "", []),
                "insight": (
                    f"The selected month highlights the five pools with the highest observed ODR."
                    if odr_options
                    else "ODR month selection is unavailable because aligned BC/TC monthly comparisons were not found."
                ),
            }
        ] if "odr_top5_by_month" in requested_features else [])
        + ([
            {
                "key": "comparison_table",
                "tab": "comparison",
                "title": "Top Segment Comparison",
                "description": "Side-by-side ranked comparison of the strongest visible segments.",
                "chart_type": "table",
                "rows": [
                    {"Segment": item["label"], "Exposure": _format_value(item["value"]), "Share": f"{item['share'] * 100:.1f}%"}
                    for item in exposure_items[:10]
                ],
                "insight": "Use this comparison table to inspect the largest contributors behind the dominant-sheet distribution." if exposure_items else "No ranked exposure rows were available for comparison.",
            }
        ] if "comparison" in tab_keys else [])
        + ([
            {
                "key": "detail_table",
                "tab": "drilldown",
                "title": "Detail Table",
                "description": "Focused item-level detail surfaced from the dominant sheet.",
                "chart_type": "table",
                "rows": [
                    {"Item": item["label"], "Exposure": _format_value(item["value"]), "Share": f"{item['share'] * 100:.1f}%"}
                    for item in exposure_items[:12]
                ],
                "insight": "This drilldown table is generated for SME-requested item-level review." if exposure_items else "No item-level rows were available for drilldown.",
            }
        ] if "drilldown" in tab_keys else []),
        "supporting_notes": [
            "This workbook is using adaptive generic EDA because no domain-specific runtime family matched it cleanly.",
            "Charts are built from the dominant analytical sheet and sheet-level movement across the workbook.",
            "Month-select controls are enabled when requested features and aligned period comparisons are available." if controls else "No interactive control was requested for this adaptive dashboard.",
        ],
    }
    return base


def _build_generic_sheet_runtime(sheet_name: str, headers: list[str], rows: list[dict[str, Any]]) -> dict[str, Any]:
    if not rows:
        return {
            "sheet_name": sheet_name,
            "sheet_kind": "empty",
            "row_count": 0,
            "column_count": len(headers),
            "headers": headers[:12],
        }

    profiles = classify_table_fields(
        type("TableLike", (), {"headers": headers, "rows": rows})()  # type: ignore[misc]
    )
    measure_headers = [header for header in headers if profiles[header].role in {"measure", "duration"}]
    numeric_headers = [header for header in measure_headers if profiles[header].role == "measure"]
    text_headers = [header for header in headers if profiles[header].role in {"dimension", "code", "identifier"}]

    if not measure_headers:
        text_items: list[str] = []
        for row in rows[:8]:
            values = [str(value).strip() for value in row.values() if value not in (None, "")]
            if values:
                text_items.append(" — ".join(values[:2]))
        return {
            "sheet_name": sheet_name,
            "sheet_kind": "text_reference",
            "row_count": len(rows),
            "column_count": len(headers),
            "headers": headers[:12],
            "text_items": text_items[:6],
        }

    dimension_header = text_headers[0] if text_headers else headers[0]
    measure_totals: list[dict[str, Any]] = []
    duration_totals: list[dict[str, Any]] = []
    for header in measure_headers:
        total = sum(coerce_measure_value(row.get(header), profiles[header]) or 0.0 for row in rows)
        target = duration_totals if profiles[header].role == "duration" else measure_totals
        label = f"{header} (hrs)" if profiles[header].role == "duration" else header
        target.append({"label": label, "total": total})
    measure_totals.sort(key=lambda item: item["total"], reverse=True)
    duration_totals.sort(key=lambda item: item["total"], reverse=True)

    segment_totals_by_label: dict[str, float] = {}
    for row in rows:
        label = str(row.get(dimension_header) or "Unknown").strip() or "Unknown"
        total = sum(coerce_measure_value(row.get(header), profiles[header]) or 0.0 for header in measure_headers)
        segment_totals_by_label[label] = segment_totals_by_label.get(label, 0.0) + total
    top_segments = [
        {"label": label, "total": total}
        for label, total in sorted(segment_totals_by_label.items(), key=lambda item: item[1], reverse=True)
    ]

    grand_total = sum(item["total"] for item in measure_totals)
    return {
        "sheet_name": sheet_name,
        "sheet_kind": "distribution",
        "row_count": len(rows),
        "column_count": len(headers),
        "headers": headers[:12],
        "dimension_header": dimension_header,
        "measure_count": len(measure_headers),
        "grand_total": grand_total,
        "measure_totals": measure_totals[:8],
        "duration_totals": duration_totals[:6],
        "top_segments": top_segments[:8],
        "segment_totals": top_segments,
        "field_roles": [
            {"label": header, "role": profiles[header].role, "data_kind": profiles[header].data_kind}
            for header in headers[:16]
        ],
    }


def build_generic_workbook_runtime(path: Path) -> dict[str, Any]:
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    sheet_summaries: list[dict[str, Any]] = []
    total_rows = 0
    total_numeric_measures = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row, header_values = _detect_header_row(ws)
        if header_row is None:
            continue
        included_columns, headers = unique_headers(header_values)
        rows: list[dict[str, Any]] = []
        for values in ws.iter_rows(
            min_row=header_row + 1,
            max_row=ws.max_row,
            min_col=1,
            max_col=ws.max_column,
            values_only=True,
        ):
            row = {
                headers[pos]: values[col_idx] if col_idx < len(values) else None
                for pos, col_idx in enumerate(included_columns)
            }
            if any(value not in (None, "") for value in row.values()):
                rows.append(row)
        summary = _build_generic_sheet_runtime(sheet_name, headers, rows)
        total_rows += int(summary.get("row_count") or 0)
        total_numeric_measures += int(summary.get("measure_count") or 0)
        sheet_summaries.append(summary)

    comparison_index: dict[str, list[dict[str, Any]]] = {}
    text_reference_items: list[str] = []
    for summary in sheet_summaries:
        if summary.get("sheet_kind") == "text_reference":
            text_reference_items.extend(summary.get("text_items") or [])
        prefix, period = _sheet_group(str(summary.get("sheet_name") or ""))
        if not prefix or not period or summary.get("sheet_kind") != "distribution":
            continue
        comparison_index.setdefault(period, []).append(
            {
                "series": prefix,
                "sheet_name": summary["sheet_name"],
                "grand_total": summary.get("grand_total") or 0.0,
            }
        )

    comparison_groups: list[dict[str, Any]] = []
    for period, items in comparison_index.items():
        if len(items) < 2:
            continue
        ranked = sorted(items, key=lambda item: item["grand_total"], reverse=True)
        leader = ranked[0]
        group_summary = {
            "group_label": period,
            "series_totals": ranked,
            "leading_series": leader["series"],
            "leading_total": leader["grand_total"],
        }
        tc_item = next((item for item in items if str(item["series"]).upper() == "TC"), None)
        bc_item = next((item for item in items if str(item["series"]).upper() == "BC"), None)
        if tc_item and bc_item:
            tc_sheet = next((summary for summary in sheet_summaries if summary.get("sheet_name") == tc_item["sheet_name"]), None)
            bc_sheet = next((summary for summary in sheet_summaries if summary.get("sheet_name") == bc_item["sheet_name"]), None)
            tc_segments = {
                str(item.get("label") or "Unknown"): float(item.get("total") or 0.0)
                for item in (tc_sheet or {}).get("segment_totals", [])
            }
            bc_segments = {
                str(item.get("label") or "Unknown"): float(item.get("total") or 0.0)
                for item in (bc_sheet or {}).get("segment_totals", [])
            }
            matched_labels = sorted(set(tc_segments) & set(bc_segments))
            ratios: list[dict[str, Any]] = []
            for label in matched_labels:
                total_count = tc_segments.get(label, 0.0)
                if total_count <= 0:
                    continue
                bad_count = bc_segments.get(label, 0.0)
                ratios.append(
                    {
                        "label": label,
                        "ratio": bad_count / total_count,
                        "bad_count": bad_count,
                        "total_count": total_count,
                    }
                )
            ratios.sort(key=lambda item: item["ratio"])
            if ratios:
                group_summary["rate_basis"] = "BC/TC"
                group_summary["matched_pool_count"] = len(ratios)
                group_summary["unmatched_tc_pool_count"] = len(set(tc_segments) - set(bc_segments))
                group_summary["unmatched_bc_pool_count"] = len(set(bc_segments) - set(tc_segments))
                group_summary["lowest_rate_segments"] = ratios[:5]
                group_summary["highest_rate_segments"] = list(reversed(ratios[-5:]))
        comparison_groups.append(group_summary)
    comparison_groups.sort(key=lambda item: str(item["group_label"]))

    return {
        "total_sheets": len(sheet_summaries),
        "tabular_sheet_count": sum(1 for item in sheet_summaries if item.get("sheet_kind") == "distribution"),
        "reference_sheet_count": sum(1 for item in sheet_summaries if item.get("sheet_kind") == "text_reference"),
        "total_rows": total_rows,
        "numeric_measure_count": total_numeric_measures,
        "comparison_group_count": len(comparison_groups),
        "sheet_summaries": sheet_summaries,
        "comparison_groups": comparison_groups[:8],
        "text_reference_items": text_reference_items[:8],
    }


def build_status_pipeline_runtime(path: Path) -> dict[str, Any]:
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [str(ws.cell(1, col).value or "").strip() for col in range(1, ws.max_column + 1)]
    normalized = {header: idx + 1 for idx, header in enumerate(headers) if header}

    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        row = {header: ws.cell(r, col_idx).value for header, col_idx in normalized.items()}
        if not any(value not in (None, "") for value in row.values()):
            continue
        rows.append(row)

    status_counts: Counter[str] = Counter()
    return_type_counts: Counter[str] = Counter()
    client_type_counts: Counter[str] = Counter()
    note_rows: list[dict[str, Any]] = []
    open_rows: list[dict[str, Any]] = []
    stale_rows: list[dict[str, Any]] = []
    today = date.today()

    for row in rows:
        status = _normalize_status(str(row.get("Return Status") or "Unknown"))
        status_counts[status] += 1
        return_type = str(row.get("Return Type") or "Unknown").strip() or "Unknown"
        return_type_counts[return_type] += 1
        client_type = str(row.get("Client Type") or "Unknown").strip() or "Unknown"
        client_type_counts[client_type] += 1

        assigned_at = _parse_datetime(row.get("Assigned on"))
        completed_at = _parse_datetime(row.get("Completed on"))
        age_days = (today - assigned_at.date()).days if assigned_at else None
        open_item = completed_at is None and status.lower() != "completed"

        queue_row = {
            "tax_payer_name": str(row.get("Tax Payer Name") or "").strip(),
            "return_code": str(row.get("Return Code") or "").strip(),
            "return_type": return_type,
            "return_status": status,
            "client_type": client_type,
            "assigned_on": assigned_at.isoformat(sep=" ") if assigned_at else None,
            "completed_on": completed_at.isoformat(sep=" ") if completed_at else None,
            "age_days": age_days,
            "cpa_notes": str(row.get("CPA Notes") or "").strip(),
            "gkm_notes": str(row.get("GKM Notes") or "").strip(),
        }
        if open_item:
            open_rows.append(queue_row)
            if age_days is not None and age_days >= 3:
                stale_rows.append(queue_row)
        if queue_row["cpa_notes"] not in ("", "_x000D_") or queue_row["gkm_notes"] not in ("", "_x000D_"):
            note_rows.append(queue_row)

    total_returns = len(rows)
    completed_returns = status_counts.get("Completed", 0)
    runtime = {
        "total_returns": total_returns,
        "completed_returns": completed_returns,
        "open_returns": total_returns - completed_returns,
        "awaiting_answers": status_counts.get("Awaiting Answers", 0),
        "under_review": status_counts.get("Under Review", 0),
        "in_process": status_counts.get("In-Process", 0) + status_counts.get("In Process", 0),
        "ready_for_preparation": status_counts.get("Ready for preparation", 0),
        "status_counts": [{"label": key, "count": value} for key, value in status_counts.most_common()],
        "return_type_counts": [{"label": key, "count": value} for key, value in return_type_counts.most_common()],
        "client_type_counts": [{"label": key, "count": value} for key, value in client_type_counts.most_common()],
        "open_queue": open_rows,
        "stale_items": stale_rows,
        "note_rows": note_rows,
    }
    return runtime


def build_product_master_runtime(path: Path) -> dict[str, Any]:
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [str(ws.cell(1, col).value or "").strip() for col in range(1, ws.max_column + 1)]
    normalized = {header: idx + 1 for idx, header in enumerate(headers) if header}

    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column, values_only=True):
        row = {header: values[col_idx - 1] for header, col_idx in normalized.items() if col_idx - 1 < len(values)}
        if not any(value not in (None, "") for value in row.values()):
            continue
        rows.append(row)

    product_type_counts: Counter[str] = Counter()
    uom_counts: Counter[str] = Counter()
    category_counts: Counter[str] = Counter()
    quality_gaps: list[dict[str, Any]] = []
    catalog_rows: list[dict[str, Any]] = []

    for row in rows:
        product_id = str(row.get("Product - ID") or "").strip()
        description = str(row.get("Product Description") or "").strip()
        product_type = str(row.get("Product Type") or "Unknown").strip() or "Unknown"
        uom = str(row.get("Base UoM") or "Unknown").strip() or "Unknown"
        category = str(row.get("Product Category") or "Unknown").strip() or "Unknown"

        product_type_counts[product_type] += 1
        uom_counts[uom] += 1
        category_counts[category] += 1
        catalog_rows.append(
            {
                "product_id": product_id,
                "description": description,
                "product_type": product_type,
                "base_uom": uom,
                "category": category,
                "hsn_code": str(row.get("HSN Code") or "").strip(),
            }
        )
        missing = [label for label, value in [("Product ID", product_id), ("Description", description), ("Product Type", product_type), ("Base UoM", uom)] if value in ("", "Unknown")]
        if missing:
            quality_gaps.append(
                {
                    "product_id": product_id or "—",
                    "description": description or "—",
                    "missing_fields": missing,
                }
            )

    return {
        "total_products": len(catalog_rows),
        "product_type_count": len(product_type_counts),
        "uom_count": len(uom_counts),
        "category_count": len(category_counts),
        "product_type_counts": [{"label": key, "count": value} for key, value in product_type_counts.most_common()],
        "uom_counts": [{"label": key, "count": value} for key, value in uom_counts.most_common()],
        "category_counts": [{"label": key, "count": value} for key, value in category_counts.most_common(12)],
        "catalog_rows": catalog_rows[:250],
        "quality_gaps": quality_gaps[:50],
    }


def get_dashboard_runtime_payload(
    db: Session,
    snapshot: models.Snapshot,
    workbook_type: str,
    dashboard_config: dict[str, Any] | None = None,
) -> dict[str, Any] | None:
    path = resolve_snapshot_workbook_path(snapshot)
    if path is None:
        return None
    if workbook_type == "client_status_report_v1":
        return build_status_pipeline_runtime(path)
    if workbook_type == "product_master_v1":
        return build_product_master_runtime(path)
    if workbook_type == "generic_workbook_v1":
        if isinstance(dashboard_config, dict) and (
            dashboard_config.get("adaptive_dashboard_enabled")
            or str(dashboard_config.get("layout_template") or "") == "adaptive_semantic"
        ):
            return build_adaptive_generic_runtime(snapshot, dashboard_config)
        return build_generic_workbook_runtime(path)
    return None
